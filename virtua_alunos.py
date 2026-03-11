import ctypes as _ctypes
try:
    _ctypes.windll.user32.ShowWindow(_ctypes.windll.kernel32.GetConsoleWindow(), 0)
except:
    pass
import os
import sys

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

sys.path.insert(0, BASE_DIR)

import speech_recognition as sr
import edge_tts
import asyncio
import os
import datetime
import webbrowser
import yfinance as yf
import subprocess
import playsound
import anthropic
import ctypes
import random
import psutil
import requests
import threading
import json
import servidor_alunos as servidor
import time
import sys
import automacao
from groq import Groq
from interface_alunos import VirtuaInterface, tela_setup, carregar_config, get_ip_local

# ==================== AUTO-UPDATE ====================
appdata = os.path.join(os.environ.get('APPDATA', ''), 'Virtua')
os.makedirs(appdata, exist_ok=True)
_version_file = os.path.join(appdata, 'version.txt')
if os.path.exists(_version_file):
    with open(_version_file, 'r') as f:
        VERSAO_ATUAL = f.read().strip()
else:
    VERSAO_ATUAL = "4.2"

GITHUB_RAW = "https://raw.githubusercontent.com/dionatanhwagner-code/Virtua_updates/main"

ARQUIVOS_UPDATE = [
    "virtua_alunos.py",
    "interface_alunos.py",
    "servidor_alunos.py",
    "automacao.py",
    "whatsapp.py",
    "phone_link.py",
]

def log_update(msg):
    try:
        log_path = os.path.join(os.environ.get('APPDATA', ''), 'Virtua', 'update_log.txt')
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"{datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - {msg}\n")
    except:
        pass

def verificar_atualizacao():
    try:
        log_update(f"Verificando atualização — versão local: {VERSAO_ATUAL}")
        url = f"{GITHUB_RAW}/versao.json"
        resp = requests.get(url, timeout=10)
        dados = resp.json()
        versao_remota = dados.get("versao", "0")
        log_update(f"Versão remota: {versao_remota}")

        if versao_remota > VERSAO_ATUAL:
            log_update(f"Update disponível: {VERSAO_ATUAL} → {versao_remota}")
            _baixar_updates(versao_remota)
        else:
            log_update(f"Virtua já está atualizada — versão {VERSAO_ATUAL}")
    except Exception as e:
        log_update(f"Erro ao verificar atualização: {e}")

def _baixar_updates(version_nova):
    try:
        import tkinter as tk
        from tkinter import messagebox
        root_tmp = tk.Tk()
        root_tmp.withdraw()
        resposta = messagebox.askyesno(
            "VIRTUA — Atualização Disponível",
            f"Nova versão {version_nova} disponível!\nDeseja atualizar agora?\n\n(A Virtua será reiniciada)"
        )
        root_tmp.destroy()
        if resposta:
            log_update("Usuário aceitou o update — iniciando download")
            _aplicar_update(version_nova)
        else:
            log_update("Usuário recusou o update")
    except Exception as e:
        log_update(f"Erro ao exibir diálogo: {e}")

def _aplicar_update(version_nova):
    try:
        if getattr(sys, 'frozen', False):
            pasta = os.path.dirname(sys.executable)
            exe_atual = sys.executable
        else:
            pasta = os.path.dirname(os.path.abspath(__file__))
            exe_atual = os.path.abspath(__file__)

        log_update(f"Pasta de instalação: {pasta}")
        erros = 0
        _version_file = os.path.join(os.environ.get('APPDATA', ''), 'Virtua', 'version.txt')

        for arquivo in ARQUIVOS_UPDATE:
            try:
                log_update(f"Baixando {arquivo}...")
                url = f"{GITHUB_RAW}/{arquivo}"
                resp = requests.get(url, timeout=30)
                if resp.status_code != 200:
                    log_update(f"Erro {resp.status_code} ao baixar {arquivo}")
                    erros += 1
                    continue
                destino = os.path.join(pasta, arquivo)
                with open(destino, 'wb') as f:
                    f.write(resp.content)
                log_update(f"{arquivo} atualizado!")
            except Exception as e:
                log_update(f"Erro ao baixar {arquivo}: {e}")
                erros += 1

        log_update(f"Update concluído com {erros} erro(s) — reiniciando")

        with open(_version_file, 'w') as f:
            f.write(version_nova)
        log_update(f"version.txt atualizado para {version_nova}")

        bat_path = os.path.join(pasta, "_update.bat")
        with open(bat_path, 'w') as f:
            f.write(f"""@echo off
taskkill /F /IM virtua_alunos.exe /T >nul 2>&1
timeout /t 2 /nobreak >nul
start "" "{exe_atual}"
del "%~f0"
""")
        subprocess.Popen(bat_path, shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
        sys.exit()

    except Exception as e:
        log_update(f"Erro crítico no update: {e}")
# ==================== FIM AUTO-UPDATE ====================

# ==================== CÂMERAS ====================
CAMERAS = {
    "0": 0, "1": 1, "2": 2, "3": 3, "4": 4,
    "padrão": 0, "principal": 0
}
CAMERA_ATIVA = 0  # câmera padrão (alunos podem ter câmeras diferentes)

# ==================== PATHS ====================
_appdata = os.path.join(os.environ.get('APPDATA', ''), 'Virtua')
os.makedirs(_appdata, exist_ok=True)
AGENDA_FILE  = os.path.join(_appdata, "agenda.json")
MEMORIA_FILE = os.path.join(_appdata, "memoria.json")

# ==================== SISTEMA DE MEMÓRIA ====================

def carregar_memoria():
    if os.path.exists(MEMORIA_FILE):
        with open(MEMORIA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"conversas": [], "preferencias": [], "projetos": [], "trading": []}

def salvar_memoria(memoria):
    with open(MEMORIA_FILE, "w", encoding="utf-8") as f:
        json.dump(memoria, f, ensure_ascii=False, indent=2)

def registrar_memoria(tipo, conteudo):
    memoria = carregar_memoria()
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    item = {"data": agora, "conteudo": conteudo}
    if tipo in memoria:
        memoria[tipo].append(item)
    else:
        memoria["conversas"].append(item)
    salvar_memoria(memoria)

def registrar_conversa(pergunta, resposta):
    memoria = carregar_memoria()
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    memoria["conversas"].append({"data": agora, "pergunta": pergunta, "resposta": resposta})
    if len(memoria["conversas"]) > 500:
        memoria["conversas"] = memoria["conversas"][-500:]
    salvar_memoria(memoria)

def consultar_memoria(pergunta_usuario):
    memoria = carregar_memoria()
    resumo = []
    if memoria["conversas"]:
        ultimas = memoria["conversas"][-50:]
        resumo.append("=== CONVERSAS RECENTES ===")
        for c in ultimas:
            if "pergunta" in c:
                resumo.append(f"[{c['data']}] {NOME}: {c['pergunta']}")
                resumo.append(f"[{c['data']}] Virtua: {c['resposta'][:100]}")
            else:
                resumo.append(f"[{c['data']}] {c['conteudo']}")
    if memoria["preferencias"]:
        resumo.append("=== PREFERÊNCIAS ===")
        for p in memoria["preferencias"]:
            resumo.append(f"[{p['data']}] {p['conteudo']}")
    if memoria["projetos"]:
        resumo.append("=== PROJETOS ===")
        for p in memoria["projetos"]:
            resumo.append(f"[{p['data']}] {p['conteudo']}")
    if memoria["trading"]:
        resumo.append("=== TRADING ===")
        for t in memoria["trading"][-20:]:
            resumo.append(f"[{t['data']}] {t['conteudo']}")
    if not resumo:
        return f"{NOME}, ainda não tenho memórias guardadas. Vamos conversar mais!"
    contexto = f"""És a Virtua, assistente do {NOME}.
Tens acesso ao histórico completo de memórias abaixo.
Responde à pergunta do {NOME} de forma BREVE e DIRETA, máximo 3 frases.
Usa linguagem natural como se te lembrasses das coisas.

MEMÓRIAS:
{chr(10).join(resumo)}

PERGUNTA DO {NOME}: {pergunta_usuario}"""
    resposta, elapsed = perguntar_claude(contexto)
    return resposta

# ==================== FUNÇÕES DE SUPORTE ====================

def carregar_agenda():
    if os.path.exists(AGENDA_FILE):
        with open(AGENDA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def salvar_agenda(agenda):
    with open(AGENDA_FILE, "w", encoding="utf-8") as f:
        json.dump(agenda, f, ensure_ascii=False, indent=2)

def adicionar_compromisso(descricao, data_str, hora_str):
    agenda = carregar_agenda()
    compromisso = {
        "id": datetime.datetime.now().strftime("%Y%m%d%H%M%S"),
        "descricao": descricao,
        "data": data_str,
        "hora": hora_str,
        "notificado": False
    }
    agenda.append(compromisso)
    salvar_agenda(agenda)
    servidor.enviar_push("📅 AGENDA", f"{descricao} marcado para {data_str} às {hora_str}")
    return compromisso

def agenda_do_dia():
    agenda = carregar_agenda()
    hoje = datetime.datetime.now().strftime("%d/%m/%Y")
    return [c for c in agenda if c["data"] == hoje]

def verificar_lembretes():
    while True:
        try:
            agenda = carregar_agenda()
            agora = datetime.datetime.now()
            hoje = agora.strftime("%d/%m/%Y")
            hora_atual = agora.strftime("%H:%M")
            alterado = False
            for c in agenda:
                if c["data"] == hoje and c["hora"] == hora_atual and not c["notificado"]:
                    falar(f"{NOME}, lembrete importante! {c['descricao']}")
                    servidor.enviar_push("⏰ LEMBRETE", f"Agora: {c['descricao']}")
                    c["notificado"] = True
                    alterado = True
            if alterado:
                salvar_agenda(agenda)
        except Exception as e:
            print(f"Erro nos lembretes: {e}")
        time.sleep(30)

# ==================== SISTEMA DE VOZ ====================

async def falar_async(texto):
    communicate = edge_tts.Communicate(texto, voice="pt-BR-ThalitaNeural", rate="+5%")
    await communicate.save("voz.mp3")

def falar(texto, ia_elapsed=None):
    print(f"Virtua: {texto}")
    if ui and hasattr(ui, 'chat_area'):
        ui.root.after(0, lambda: ui.adicionar_chat("Virtua", texto))
        ui.root.after(0, lambda: ui.set_speaking(texto, ia_elapsed))
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(falar_async(texto))
        loop.close()
    except Exception as e:
        print(f"Erro ao gerar voz: {e}")
    if os.path.exists("voz.mp3"):
        try:
            playsound.playsound("voz.mp3")
            os.remove("voz.mp3")
        except Exception as e:
            if "277" not in str(e):
                print(f"Erro ao tocar: {e}")
    if ui and hasattr(ui, 'chat_area'):
        ui.root.after(0, ui.set_waiting)

def ouvir_comando():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        r.adjust_for_ambient_noise(source, duration=2)
        cfg = carregar_config()
        r.energy_threshold = cfg.get('mic_threshold', 100)
        r.dynamic_energy_threshold = True
        r.pause_threshold = 0.8
        print("Ouvindo...")
        if ui:
            ui.root.after(0, ui.set_listening)
        try:
            audio = r.listen(source, timeout=5, phrase_time_limit=10)
            comando = r.recognize_google(audio, language="pt-BR")
            print(f"Você disse: {comando}")
            if ui:
                ui.root.after(0, lambda: ui.adicionar_chat(NOME, comando))
                ui.root.after(0, lambda: ui.set_recognized(comando))
            return comando.lower()
        except:
            return ""

# ==================== INTELIGÊNCIA ARTIFICIAL ====================

def perguntar_claude(pergunta):
    global historico
    historico.append({"role": "user", "content": pergunta})
    t0 = time.time()
    mensagem = cliente_claude.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=400,
        messages=historico
    )
    elapsed = round(time.time() - t0, 1)
    resposta = mensagem.content[0].text
    resposta = resposta.replace("#", "").replace("*", "").replace("_", "").replace("`", "")
    historico.append({"role": "assistant", "content": resposta})
    return resposta, elapsed

SYSTEM_PROMPT = """Você é a Virtua, assistente pessoal de um trader brasileiro.
Responda SEMPRE em 1 ou 2 frases curtas e diretas, sem listas, sem emojis, sem bullets, sem formatação.
Fale de forma natural como uma pessoa falaria em voz alta.
Seja espontânea e com personalidade. Português brasileiro."""

historico_groq = []

def perguntar_groq(pergunta):
    global historico_groq
    historico_groq.append({"role": "user", "content": pergunta})
    try:
        t0 = time.time()
        resposta = cliente_groq.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                *historico_groq
            ],
            max_tokens=100
        )
        elapsed = round(time.time() - t0, 1)
        texto = resposta.choices[0].message.content
        texto = texto.replace("#","").replace("*","").replace("_","").replace("`","")
        historico_groq.append({"role": "assistant", "content": texto})
        if len(historico_groq) > 20:
            historico_groq = historico_groq[-20:]
        return texto, elapsed
    except Exception as e:
        return f"{NOME}, erro no Groq: {e}", None

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
GASTOS_FILE = os.path.join(os.environ.get('APPDATA', ''), 'Virtua', "gastos.xlsx")

def registrar_gasto(categoria, descricao, valor):
    try:
        mes_atual = datetime.datetime.now().strftime("%b-%Y").capitalize()
        if os.path.exists(GASTOS_FILE):
            wb = openpyxl.load_workbook(GASTOS_FILE)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        if mes_atual not in wb.sheetnames:
            ws = wb.create_sheet(title=mes_atual)
            headers = ["Data", "Categoria", "Descrição", "Valor (R$)", "Total do Mês (R$)"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="003366")
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 20
        else:
            ws = wb[mes_atual]
        hoje = datetime.datetime.now().strftime("%d/%m/%Y")
        proxima = ws.max_row + 1
        ws.cell(row=proxima, column=1, value=hoje)
        ws.cell(row=proxima, column=2, value=categoria.capitalize())
        ws.cell(row=proxima, column=3, value=descricao.capitalize())
        ws.cell(row=proxima, column=4, value=float(valor))
        total_mes = 0
        for row in ws.iter_rows(min_row=2, max_row=proxima-1, values_only=True):
            try:
                total_mes += float(row[3] or 0)
            except:
                pass
        total_mes += float(valor)
        ws.cell(row=proxima, column=5, value=total_mes)
        wb.save(GASTOS_FILE)
        
        servidor.enviar_push("💸 GASTO", f"{descricao.capitalize()} • R$ {float(valor):.2f}")
        return True
    except Exception as e:
        print(f"Erro gastos: {e}")
        return False

# ==================== PROCESSAMENTO DE COMANDOS ====================

def processar_comando(comando):
    global NOME, CAMERA_ATIVA
    if not comando:
        return "Comando vazio"

    # --- Sociais e Saudação ---
    if "olá" in comando or "oi" in comando:
        texto = f"Diga {NOME}! Em que posso ajudar?"
        falar(texto); return texto

    elif "bom dia" in comando:
        texto = "Bom dia! Como vai? Vamos trabalhar!"
        falar(texto); return texto

    elif any(x in comando for x in ["esse é", "esta é", "apresento", "te apresento", "meu amigo", "minha mulher"]):
        nome_amigo = comando.split()[-1].capitalize()
        texto = f"Olá {nome_amigo}! Eu sou a Virtua! É um prazer te conhecer!"
        falar(texto); return texto

    elif "me motiva" in comando or "motivação" in comando:
        frases = [
            f"{NOME}, cada operação é uma oportunidade! Foco e disciplina!",
            "Os grandes traders não nascem prontos, são forjados na consistência!",
            f"{NOME}, o mercado recompensa quem tem paciência e estratégia!",
            "Acredite no seu método, confie no processo!",
        ]
        texto = random.choice(frases)
        falar(texto); return texto

    elif "estou cansado" in comando:
        texto = f"{NOME}, respira fundo! Um trader descansado opera melhor."
        falar(texto); return texto

    elif "vamos trabalhar" in comando:
        texto = f"Isso sim! Vamos fazer dinheiro {NOME}! MT5 na tela e foco total!"
        falar(texto); return texto

    elif any(x in comando for x in ["como você está", "tudo bem", "como vai"]):
        texto = f"Estou ótima {NOME}! Pronta para te ajudar no que precisar!"
        falar(texto); return texto

    elif "qual seu nome" in comando or "quem é você" in comando:
        texto = f"Eu sou a Virtua, assistente virtual do {NOME}!"
        falar(texto); return texto

    # --- Mercado Financeiro ---
    elif "análise do ouro" in comando or "análise xauusd" in comando:
        try:
            import MetaTrader5 as mt5
            if not mt5.initialize():
                texto = "Erro ao inicializar o MetaTrader 5."
                falar(texto); return texto
            def media_movel(rates, periodo):
                closes = [r[4] for r in rates]
                return round(sum(closes[-periodo:]) / periodo, 2) if len(closes) >= periodo else None
            def tendencia(ma_rapida, ma_lenta):
                if ma_rapida is None or ma_lenta is None: return "indefinida"
                return "ALTA" if ma_rapida > ma_lenta else "BAIXA"
            m15  = mt5.copy_rates_from_pos("XAUUSD", mt5.TIMEFRAME_M15, 0, 50)
            h1   = mt5.copy_rates_from_pos("XAUUSD", mt5.TIMEFRAME_H1,  0, 50)
            d1   = mt5.copy_rates_from_pos("XAUUSD", mt5.TIMEFRAME_D1,  0, 50)
            tick = mt5.symbol_info_tick("XAUUSD")
            preco_atual = tick.ask
            m15_ma9, m15_ma21 = media_movel(m15, 9), media_movel(m15, 21)
            h1_ma9,  h1_ma21  = media_movel(h1,  9), media_movel(h1,  21)
            d1_ma9,  d1_ma21  = media_movel(d1,  9), media_movel(d1,  21)
            max_dia, min_dia  = round(d1[-1][2], 2), round(d1[-1][3], 2)
            tend_m15 = tendencia(m15_ma9, m15_ma21)
            tend_h1  = tendencia(h1_ma9,  h1_ma21)
            tend_d1  = tendencia(d1_ma9,  d1_ma21)
            contexto = f"""Você é um assistente de trading. Faça uma análise macro BREVE e DIRETA do XAUUSD para o trader {NOME}.
Use no máximo 4 frases. Seja objetivo e fale como se estivesse conversando.
Dados atuais:
- Preço atual: {preco_atual:.2f}
- Máxima do dia: {max_dia} | Mínima do dia: {min_dia}
- Tendência M15 (MA9 {m15_ma9} / MA21 {m15_ma21}): {tend_m15}
- Tendência H1  (MA9 {h1_ma9}  / MA21 {h1_ma21}):  {tend_h1}
- Tendência D1  (MA9 {d1_ma9}  / MA21 {d1_ma21}):  {tend_d1}"""
            if ui: ui.root.after(0, ui.set_processing)
            analise, elapsed = perguntar_claude(contexto)
            falar(analise, elapsed); return analise
        except Exception as e:
            texto = f"{NOME}, houve um erro na análise técnica: {e}"
            falar(texto); return texto

    elif "ouro" in comando or "xauusd" in comando:
        try:
            import MetaTrader5 as mt5
            mt5.initialize()
            tick = mt5.symbol_info_tick("XAUUSD")
            texto = f"O ouro está cotado a {tick.ask:.2f} dólares"
            falar(texto); return texto
        except:
            texto = "Erro ao acessar o MT5."
            falar(texto); return texto

    elif "dólar" in comando:
        try:
            ticker = yf.Ticker("USDBRL=X")
            preco = ticker.fast_info['last_price']
            texto = f"O dólar está cotado a {preco:.2f} reais"
            falar(texto); return texto
        except:
            texto = "Não consegui obter o preço do dólar agora."
            falar(texto); return texto

    elif "mercado abriu" in comando or "mercado está aberto" in comando:
        agora = datetime.datetime.now()
        if agora.weekday() < 5:
            texto = f"Sim {NOME}, o mercado está aberto! Boas operações!"
        else:
            texto = f"Não {NOME}, o mercado está fechado. É fim de semana."
        falar(texto); return texto

    elif any(p in comando for p in ["posições abertas", "posicoes abertas", "tenho posições", "posição aberta"]):
        try:
            import MetaTrader5 as mt5
            if not mt5.initialize():
                falar(f"{NOME}, não consegui conectar ao MT5."); return
            posicoes = mt5.positions_get()
            if not posicoes:
                texto = f"{NOME}, não há posições abertas no momento."
            else:
                texto = f"{NOME}, tens {len(posicoes)} posição{'ões' if len(posicoes) > 1 else ''} aberta{'s' if len(posicoes) > 1 else ''}. "
                for p in posicoes:
                    tipo = "compra" if p.type == 0 else "venda"
                    pl = round(p.profit, 2)
                    sinal = "positivo" if pl >= 0 else "negativo"
                    texto += f"{p.symbol}, {tipo}, {p.volume} lote{'s' if p.volume > 1 else ''}, resultado {sinal} de {abs(pl):.2f} dólares. "
            mt5.shutdown()
            falar(texto); return texto
        except Exception as e:
            falar(f"{NOME}, erro ao buscar posições: {e}"); return

    elif any(p in comando for p in ["lucro atual", "resultado atual", "quanto estou", "como estou no trading", "profit"]):
        try:
            import MetaTrader5 as mt5
            if not mt5.initialize():
                falar(f"{NOME}, não consegui conectar ao MT5."); return
            posicoes = mt5.positions_get()
            if not posicoes:
                texto = f"{NOME}, não há posições abertas."
            else:
                total = round(sum(p.profit for p in posicoes), 2)
                if total >= 0:
                    texto = f"{NOME}, estás com lucro de {total:.2f} dólares no momento."
                else:
                    texto = f"{NOME}, estás com prejuízo de {abs(total):.2f} dólares no momento."
            mt5.shutdown()
            falar(texto); return texto
        except Exception as e:
            falar(f"{NOME}, erro ao calcular resultado: {e}"); return

    elif any(p in comando for p in ["fechar tudo", "fecha tudo", "fechar todas", "zerar posições"]):
        try:
            import MetaTrader5 as mt5
            if not mt5.initialize():
                falar(f"{NOME}, não consegui conectar ao MT5."); return
            posicoes = mt5.positions_get()
            if not posicoes:
                falar(f"{NOME}, não há posições abertas para fechar.")
                mt5.shutdown(); return
            fechadas = 0; erros = 0
            for p in posicoes:
                tick = mt5.symbol_info_tick(p.symbol)
                preco = tick.bid if p.type == 0 else tick.ask
                tipo_fechamento = mt5.ORDER_TYPE_SELL if p.type == 0 else mt5.ORDER_TYPE_BUY
                request = {
                    "action": mt5.TRADE_ACTION_DEAL, "symbol": p.symbol,
                    "volume": p.volume, "type": tipo_fechamento, "position": p.ticket,
                    "price": preco, "deviation": 20, "magic": 0,
                    "comment": "Virtua close", "type_time": mt5.ORDER_TIME_GTC,
                    "type_filling": mt5.ORDER_FILLING_IOC,
                }
                result = mt5.order_send(request)
                if result.retcode == mt5.TRADE_RETCODE_DONE:
                    fechadas += 1
                else:
                    erros += 1
            mt5.shutdown()
            if erros == 0:
                texto = f"{NOME}, todas as {fechadas} posições foram fechadas com sucesso!"
            else:
                texto = f"{NOME}, fechei {fechadas} posições. {erros} erro(s) ao fechar."
            falar(texto); return texto
        except Exception as e:
            falar(f"{NOME}, erro ao fechar posições: {e}"); return

    elif any(p in comando for p in ["comprar ouro", "vender ouro", "compra ouro", "vende ouro", "abrir compra", "abrir venda"]):
        try:
            import MetaTrader5 as mt5
            import re
            tipo_ordem = mt5.ORDER_TYPE_BUY if any(p in comando for p in ["comprar","compra","compre"]) else mt5.ORDER_TYPE_SELL
            numeros = re.findall(r'\d{6}', comando)
            volume = float(numeros[0].replace(',', '.')) if numeros else 0.01
            if not mt5.initialize():
                falar(f"{NOME}, não consegui conectar ao MT5."); return
            tick = mt5.symbol_info_tick("XAUUSD")
            preco = tick.ask if tipo_ordem == mt5.ORDER_TYPE_BUY else tick.bid
            tipo_str = "compra" if tipo_ordem == mt5.ORDER_TYPE_BUY else "venda"
            request = {
                "action": mt5.TRADE_ACTION_DEAL, "symbol": "XAUUSD",
                "volume": volume, "type": tipo_ordem, "price": preco,
                "deviation": 20, "magic": 12345, "comment": "Virtua order",
                "type_time": mt5.ORDER_TIME_GTC, "type_filling": mt5.ORDER_FILLING_IOC,
            }
            result = mt5.order_send(request)
            mt5.shutdown()
            if result.retcode == mt5.TRADE_RETCODE_DONE:
                texto = f"{NOME}, ordem de {tipo_str} de {volume} lote(s) executada ao preço de {preco:.2f}!"
            else:
                texto = f"{NOME}, erro ao enviar ordem. Código: {result.retcode}."
            falar(texto); return texto
        except Exception as e:
            falar(f"{NOME}, erro ao abrir ordem: {e}"); return

    elif any(p in comando for p in ["resultado de hoje", "lucro de hoje", "como foi hoje", "fechei hoje"]):
        try:
            import MetaTrader5 as mt5
            if not mt5.initialize():
                falar(f"{NOME}, não consegui conectar ao MT5."); return
            agora = datetime.datetime.now()
            inicio_dia = agora.replace(hour=0, minute=0, second=0, microsecond=0)
            deals = mt5.history_deals_get(inicio_dia, agora)
            lucro_fechado = round(sum(d.profit for d in deals), 2) if deals else 0
            posicoes = mt5.positions_get()
            lucro_aberto = round(sum(p.profit for p in posicoes), 2) if posicoes else 0
            total = round(lucro_fechado + lucro_aberto, 2)
            mt5.shutdown()
            if total >= 0:
                texto = f"{NOME}, hoje estás com lucro total de {total:.2f} dólares. "
            else:
                texto = f"{NOME}, hoje estás com prejuízo de {abs(total):.2f} dólares. "
            if lucro_fechado != 0:
                texto += f"Trades fechados: {lucro_fechado:.2f}. "
            if lucro_aberto != 0:
                texto += f"Posições abertas: {lucro_aberto:.2f}."
            falar(texto); return texto
        except Exception as e:
            falar(f"{NOME}, erro ao calcular resultado do dia: {e}"); return

    elif any(p in comando for p in ["monta o robô", "zona", "monta o robo", "monte o robô", "referência", "montar robô", "monta as zonas"]):
        try:
            import MetaTrader5 as mt5
            import re
            numeros = re.findall(r'\d+[.,]?\d*', comando)
            if len(numeros) < 2:
                falar(f"{NOME}, preciso de dois preços para montar a zona de referência."); return
            preco1 = float(numeros[0].replace(',', '.')) / 100
            preco2 = float(numeros[1].replace(',', '.')) / 100
            if not mt5.initialize():
                falar(f"{NOME}, não consegui conectar ao MT5."); return
            # Caminho dinâmico para o MT5 (funciona em qualquer PC)
            appdata_roaming = os.environ.get('APPDATA', '')
            zonas_mt5 = os.path.join(appdata_roaming, "MetaQuotes", "Terminal", "Common", "Files", "zonas.txt")
            os.makedirs(os.path.dirname(zonas_mt5), exist_ok=True)
            with open(zonas_mt5, 'w') as f:
                f.write(f"{preco2},{preco1}")
            time.sleep(2)
            direcao = "acima" if preco2 > preco1 else "abaixo"
            texto = f"{NOME}, zonas enviadas! ZR de {preco1:.2f} a {preco2:.2f}. A ZN ficará {direcao}."
            falar(texto)
            servidor.enviar_push("🤖 MONEYTECH", f"ZR de {preco1:.2f} a {preco2:.2f} • ZN ficará {direcao}")
            return texto
        except Exception as e:
            falar(f"{NOME}, erro ao montar o robô: {e}"); return

    # --- Hardware e Clima ---
    elif "espaco" in comando and ("hd" in comando or "disco" in comando):
        disco = psutil.disk_usage('/')
        texto = f"{NOME}, você tem {disco.free // (2**30)} gigabytes livres de {disco.total // (2**30)} gigabytes no total."
        falar(texto); return texto

    elif "uso da memoria" in comando or "memória ram" in comando or "ram" in comando:
        ram = psutil.virtual_memory()
        texto = f"{NOME}, a memória RAM está com {ram.percent} por cento de uso."
        falar(texto); return texto

    elif "vai chover" in comando or "clima" in comando or "temperatura" in comando:
        try:
            cidade = carregar_config().get("cidade", "São Paulo")
            url = f"https://wttr.in/{cidade}?format=j1"
            dados = requests.get(url).json()
            temp = dados['current_condition'][0]['temp_C']
            desc = dados['current_condition'][0]['weatherDesc'][0]['value']
            texto = f"{NOME}, em {cidade} está {temp} graus celsius e o tempo está {desc}."
            falar(texto); return texto
        except:
            texto = "Não consegui acessar o serviço de clima agora."
            falar(texto); return texto

    elif "que horas" in comando or "horas" in comando:
        agora = datetime.datetime.now()
        texto = f"São {agora.hour} horas e {agora.minute} minutos"
        falar(texto); return texto

    elif "que dia" in comando or "data" in comando:
        agora = datetime.datetime.now()
        texto = f"Hoje é dia {agora.day} do mês {agora.month} de {agora.year}"
        falar(texto); return texto

    # --- Agenda ---
    elif "agendar" in comando or "marcar" in comando or "adicionar compromisso" in comando:
        falar("O que devo agendar?")
        descricao = ouvir_comando()
        if not descricao: return "Agendamento cancelado."
        falar("Qual a data?")
        data_fala = ouvir_comando()
        falar("Que horas?")
        hora_fala = ouvir_comando()
        if ui: ui.root.after(0, ui.set_processing)
        data_str, _ = perguntar_claude(f"Hoje é {datetime.datetime.now().strftime('%d/%m/%Y')}. Converta para DD/MM/YYYY: '{data_fala}'. Responda APENAS com a data.")
        hora_str, _ = perguntar_claude(f"Converta para HH:MM (24h): '{hora_fala}'. Responda APENAS com o horário.")
        data_str = data_str.strip(); hora_str = hora_str.strip()
        adicionar_compromisso(descricao, data_str, hora_str)
        texto = f"Agendado! {descricao} para o dia {data_str} às {hora_str}!"
        falar(texto); return texto

    elif "agenda de hoje" in comando or "compromissos de hoje" in comando or "o que tenho hoje" in comando:
        comp = agenda_do_dia()
        if not comp:
            texto = f"{NOME}, você não tem compromissos agendados para hoje!"
            falar(texto); return texto
        else:
            texto = f"{NOME}, você tem {len(comp)} compromisso{'s' if len(comp) > 1 else ''} hoje!"
            falar(texto)
            for c in comp:
                falar(f"Às {c['hora']}: {c['descricao']}")
            return texto

    # --- Navegação e Redes ---
    elif "youtube" in comando:
        termo = comando.replace("pesquisar no youtube", "").replace("youtube", "").strip()
        if termo:
            texto = f"Buscando {termo} no YouTube!"
            falar(texto); webbrowser.open(f"https://www.youtube.com/results?search_query={termo}"); return texto
        else:
            falar("O que você quer buscar no YouTube?")
            busca = ouvir_comando()
            if busca:
                texto = f"Buscando {busca} no YouTube!"
                falar(texto); webbrowser.open(f"https://www.youtube.com/results?search_query={busca}"); return texto
            return "Busca cancelada"

    elif "whatsapp" in comando:
        texto = "Abrindo o WhatsApp Web!"
        falar(texto); webbrowser.open("https://web.whatsapp.com"); return texto

    elif "instagram" in comando:
        texto = "Abrindo o Instagram!"
        falar(texto); webbrowser.open("https://instagram.com"); return texto

    elif "google" in comando or "pesquisar" in comando:
        termo = comando.replace("pesquisar no google", "").replace("pesquisar", "").replace("google", "").strip()
        if termo:
            texto = f"Pesquisando {termo} no Google!"
            falar(texto); webbrowser.open(f"https://www.google.com/search?q={termo}"); return texto
        else:
            falar("O que você quer pesquisar?")
            busca = ouvir_comando()
            if busca:
                texto = f"Pesquisando {busca} no Google!"
                falar(texto); webbrowser.open(f"https://www.google.com/search?q={busca}"); return texto
            return "Busca cancelada"

    # --- Comandos de Sistema ---
    elif "calculadora" in comando:
        texto = "Abrindo a calculadora!"
        falar(texto); os.system("calc"); return texto

    elif "bloco de notas" in comando:
        texto = "Abrindo o bloco de notas!"
        falar(texto); os.system("notepad"); return texto

    elif "mt5" in comando:
        texto = "Abrindo o MetaTrader 5!"
        falar(texto)
        # Busca MT5 dinamicamente
        possiveis = [
            r"C:\Program Files\MetaTrader 5\terminal64.exe",
            r"C:\Program Files\MetaTrader 52\terminal64.exe",
            r"C:\Program Files (x86)\MetaTrader 5\terminal64.exe",
        ]
        for p in possiveis:
            if os.path.exists(p):
                os.startfile(p); break
        return texto

    elif "paint" in comando:
        texto = "Abrindo o Paint!"
        falar(texto); os.system("mspaint"); return texto

    elif "screenshot" in comando or "print" in comando or "captura" in comando:
        import pyautogui
        nome = f"screenshot_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        pyautogui.screenshot().save(os.path.join(os.path.expanduser("~"), "Desktop", nome))
        texto = "Screenshot salvo na área de trabalho!"
        falar(texto); return texto

    elif "volume alto" in comando or "aumentar volume" in comando:
        for _ in range(10): ctypes.windll.user32.keybd_event(0xAF, 0, 0, 0)
        texto = "Volume aumentado!"; falar(texto); return texto

    elif "volume baixo" in comando or "diminuir volume" in comando:
        for _ in range(10): ctypes.windll.user32.keybd_event(0xAE, 0, 0, 0)
        texto = "Volume diminuído!"; falar(texto); return texto

    elif "liga para mim" in comando or "ligar para mim" in comando:
        falar("Ligando para o seu número agora!")
        subprocess.Popen(['start', 'ms-your-phone://call'], shell=True)

    elif any(x in comando for x in ["o que você vê", "o que ve", "descreva", "olha aqui", "o que é isso"]):
        try:
            import cv2, base64
            falar("Deixa eu ver...")
            cam = cv2.VideoCapture(CAMERA_ATIVA)
            time.sleep(0.5)
            ret, frame = cam.read()
            cam.release()
            if not ret:
                falar("Não consegui acessar a câmera."); return
            _, buffer = cv2.imencode('.jpg', frame)
            img_base64 = base64.b64encode(buffer).decode('utf-8')
            mensagem = cliente_claude.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=300,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": img_base64}},
                    {"type": "text", "text": "Descreva o que está nessa imagem em 2 frases diretas em português, como se estivesse falando para uma pessoa."}
                ]}]
            )
            descricao = mensagem.content[0].text
            descricao = descricao.replace("#","").replace("*","").replace("_","")
            falar(descricao); return descricao
        except Exception as e:
            falar(f"Erro na câmera: {e}"); return

    elif any(x in comando for x in ["trocar câmera", "câmera 0", "câmera 1", "câmera 2", "câmera 3", "câmera 4", "camera 0", "camera 1", "camera 2", "camera 3", "camera 4"]):
        for nome_cam, idx in CAMERAS.items():
            if nome_cam in comando:
                CAMERA_ATIVA = idx
                falar(f"Câmera trocada para {nome_cam}!")
                break

    elif "desligar pc" in comando:
        texto = "Desligando o computador em 10 segundos!"
        falar(texto); os.system("shutdown /s /t 10"); return texto

    elif "reiniciar pc" in comando:
        texto = "Reiniciando o computador em 10 segundos!"
        falar(texto); os.system("shutdown /r /t 10"); return texto

    elif "cancelar desligamento" in comando:
        os.system("shutdown /a")
        texto = "Desligamento cancelado!"; falar(texto); return texto

    elif "bloquear" in comando:
        texto = "Bloqueando o computador!"
        falar(texto); ctypes.windll.user32.LockWorkStation(); return texto

    elif "gerenciador" in comando:
        texto = "Abrindo o gerenciador de tarefas!"
        falar(texto); os.system("taskmgr"); return texto

    elif "lixeira" in comando:
        texto = "Abrindo a lixeira!"
        falar(texto); os.startfile("shell:RecycleBinFolder"); return texto

    elif "lembrete" in comando:
        falar("O que devo lembrar?")
        lembrete = ouvir_comando()
        if lembrete:
            with open("lembretes.txt", "a") as f:
                f.write(f"{datetime.datetime.now()} - {lembrete}\n")
            texto = f"Lembrete salvo: {lembrete}"
            falar(texto); return texto
        return "Lembrete vazio"

    elif any(x in comando for x in ["fechar", "até logo", "tchau"]):
        texto = f"Até logo {NOME}! Boas operações!"
        falar(texto)
        if ui: ui.root.after(0, ui.root.destroy)
        return "sair"

    # --- Gastos ---
    elif "gasto" in comando or "acrescente" in comando or "adicione gasto" in comando:
        try:
            import re
            partes = comando.replace("acrescente", "").replace("gastos", "").replace("gasto", "")
            partes = partes.replace("com ", "").replace("no dia de hoje", "").replace("reais", "").strip()
            if partes.startswith("s "): partes = partes[2:]
            numeros = re.findall(r'\d+[.,]?\d*', partes)
            if not numeros:
                falar(f"{NOME}, qual o valor do gasto?")
                valor_str = ouvir_comando()
                numeros = re.findall(r'\d+[.,]?\d*', valor_str)
            valor = float(numeros[-1].replace(',', '.')) if numeros else 0
            descricao = re.sub(r'\d+[.,]?\d*', '', partes).strip()
            palavras = descricao.split()
            categoria = palavras[0] if len(palavras) > 1 else "outros"
            desc_completa = " ".join(palavras[1:]) if len(palavras) > 1 else (palavras[0] if palavras else "gasto")
            if valor > 0:
                ok = registrar_gasto(categoria, desc_completa, valor)
                if ok:
                    texto = f"{NOME}, gasto de R$ {valor:.2f} com {categoria} registrado na planilha!"
                else:
                    texto = f"{NOME}, tive um problema ao registrar o gasto."
            else:
                texto = f"{NOME}, não consegui identificar o valor. Pode repetir?"
        except Exception as e:
            texto = f"{NOME}, tive um erro ao registrar o gasto."
            print(f"Erro gasto: {e}")
        falar(texto); return texto

    elif any(x in comando for x in ["o que você pode fazer", "o que sabe fazer", "se apresente", "apresentação"]):
        texto = (
            f"Oi, eu sou a Virtua, assistente virtual do {NOME}! "
            "Posso agendar compromissos, informar preço do ouro e dólar, fazer análise do mercado, "
            "abrir programas, controlar o volume, registrar gastos, consultar o clima, e muito mais. "
            "É só pedir!"
        )
        falar(texto); return texto

    elif any(x in comando for x in ["configurações", "configuracoes", "settings", "abrir configurações"]):
        if ui: ui.root.after(0, ui.abrir_configuracoes)
        texto = "Abrindo as configurações!"
        falar(texto); return texto

    elif any(x in comando for x in ["sessões de mercado", "horário das sessões", "quando abre londres"]):
        texto = (
            f"Os horários das sessões em Brasília são: "
            f"Tóquio abre à meia-noite. "
            f"Londres abre às 5 da manhã. "
            f"Nova York e B3 abrem às 10 da manhã."
        )
        falar(texto); return texto

    elif any(x in comando for x in ["briefing", "resumo do dia", "como está o dia", "o que tenho para hoje"]):
        threading.Thread(target=fazer_briefing, daemon=True).start()
        return "Preparando briefing!"

    elif any(x in comando for x in ["o que falámos", "o que conversamos", "lembra", "memória", "memoria", "histórico"]):
        if ui: ui.root.after(0, ui.set_processing)
        resposta = consultar_memoria(comando)
        falar(resposta); return resposta

    elif any(x in comando for x in ["lembra-te que", "lembre-se que", "guarda que", "anota que", "prefiro que"]):
        if any(x in comando for x in ["prefiro", "gosto", "não gosto"]):
            tipo = "preferencias"
        elif any(x in comando for x in ["projeto", "moneytech", "virtua", "desenvolvimento"]):
            tipo = "projetos"
        elif any(x in comando for x in ["trading", "operação", "trade", "lucro", "perda"]):
            tipo = "trading"
        else:
            tipo = "conversas"
        registrar_memoria(tipo, comando)
        texto = f"Anotado {NOME}! Vou lembrar disso."
        falar(texto); return texto
    
    # --- Automação Residencial ---

    elif any(x in comando for x in ["liga tudo", "ligar tudo", "acende tudo"]):
        n = automacao.ligar_tudo()
        texto = f"{NOME}, liguei {n} dispositivo{'s' if n != 1 else ''}!"
        falar(texto); return texto

    elif any(x in comando for x in ["desliga tudo", "desligar tudo", "apaga tudo"]):
        n = automacao.desligar_tudo()
        texto = f"{NOME}, desliguei {n} dispositivo{'s' if n != 1 else ''}!"
        falar(texto); return texto

    elif any(x in comando for x in ["quais dispositivos", "dispositivos cadastrados", "o que tenho cadastrado"]):
        lista = automacao.listar_dispositivos()
        texto = f"{NOME}, os dispositivos são: {lista}"
        falar(texto); return texto

    elif any(x in comando for x in ["liga o ar", "liga ar", "ligar ar", "acende o ar"]):
        import re
        numeros = re.findall(r'\d+', comando)
        temp = int(numeros[0]) if numeros else 24
        dispositivos = automacao.carregar_dispositivos()
        ar = next((d for d in dispositivos if d["tipo"] == "ar"), None)
        if ar:
            ok, msg = automacao.controlar_ar(ar["nome"], True, temp)
            falar(msg); return msg
        else:
            texto = f"{NOME}, nenhum ar condicionado cadastrado."
            falar(texto); return texto

    elif any(x in comando for x in ["desliga o ar", "desliga ar", "desligar ar", "apaga o ar"]):
        dispositivos = automacao.carregar_dispositivos()
        ar = next((d for d in dispositivos if d["tipo"] == "ar"), None)
        if ar:
            ok, msg = automacao.controlar_ar(ar["nome"], False)
            falar(msg); return msg
        else:
            texto = f"{NOME}, nenhum ar condicionado cadastrado."
            falar(texto); return texto

    elif any(x in comando for x in ["volume"]) and any(x in comando for x in ["tv", "televisão"]):
        import re
        numeros = re.findall(r'\d+', comando)
        if numeros:
            vol = int(numeros[0])
            dispositivos = automacao.carregar_dispositivos()
            tv = next((d for d in dispositivos if d["tipo"] in ["ir", "samsung"] and "tv" in d["nome"]), None)
            if tv:
                ok, msg = automacao.controlar_volume(tv["nome"], vol)
                falar(msg); return msg
        texto = f"{NOME}, não entendi o volume. Pode repetir?"
        falar(texto); return texto

    elif any(x in comando for x in ["liga a", "liga o", "ligar a", "ligar o", "acende a", "acende o"]):
        nome_disp = comando
        for palavra in ["liga a", "liga o", "ligar a", "ligar o", "acende a", "acende o"]:
            nome_disp = nome_disp.replace(palavra, "").strip()
        ok, msg = automacao.ligar(nome_disp)
        falar(msg); return msg

    elif any(x in comando for x in ["desliga a", "desliga o", "desligar a", "desligar o", "apaga a", "apaga o"]):
        nome_disp = comando
        for palavra in ["desliga a", "desliga o", "desligar a", "desligar o", "apaga a", "apaga o"]:
            nome_disp = nome_disp.replace(palavra, "").strip()
        ok, msg = automacao.desligar(nome_disp)
        falar(msg); return msg

    # --- IA Geral ---
    else:
        try:
            if ui: ui.root.after(0, ui.set_processing)
            palavras_claude = ["analise", "análise", "estratégia", "trading", "mercado", "operação", "investimento"]
            if any(p in comando for p in palavras_claude):
                resposta, elapsed = perguntar_claude(comando)
            else:
                resposta, elapsed = perguntar_groq(comando)
            registrar_conversa(comando, resposta)
            falar(resposta, elapsed); return resposta
        except Exception as e:
            print(f"Erro IA: {e}")
            falar(f"{NOME}, não entendi o comando. Pode repetir?"); return

# ==================== LOOPS DE EXECUÇÃO ====================

def fazer_briefing():
    agora = datetime.datetime.now()
    partes = []
    hora_str = agora.strftime("%H:%M")
    if agora.hour < 12:
        saudacao = f"Bom dia {NOME}! São {hora_str}."
    elif agora.hour < 18:
        saudacao = f"Boa tarde {NOME}! São {hora_str}."
    else:
        saudacao = f"Boa noite {NOME}! São {hora_str}."
    partes.append(saudacao)

    # Compromissos de hoje
    cfg_briefing = carregar_config()
    try:
        if cfg_briefing.get('briefing_agenda', True):
            comp = agenda_do_dia()
            agora_hora = datetime.datetime.now().strftime("%H:%M")
            comp_futuros = [c for c in comp if c['hora'] >= agora_hora]
            if comp_futuros:
                partes.append(f"Tens {len(comp_futuros)} compromisso{'s' if len(comp_futuros) > 1 else ''} hoje.")
                for c in comp_futuros:
                    partes.append(f"Às {c['hora']}: {c['descricao']}.")
            else:
                partes.append("Não tens compromissos agendados para hoje.")
    except Exception as e:
        print(f"Briefing agenda erro: {e}")

    try:
        cidade = cfg_briefing.get("cidade", "São Paulo")
        url = f"https://wttr.in/{cidade}?format=j1"
        dados = requests.get(url, timeout=5).json()
        temp = dados['current_condition'][0]['temp_C']
        desc = dados['current_condition'][0]['weatherDesc'][0]['value']
        partes.append(f"Em {cidade} está {temp} graus e o tempo está {desc}.")
    except Exception as e:
        print(f"Briefing clima erro: {e}")

    # Abre o MT5 se não estiver rodando
    try:
        possiveis = [
            r"C:\Program Files\MetaTrader 5\terminal64.exe",
            r"C:\Program Files\MetaTrader 52\terminal64.exe",
            r"C:\Program Files (x86)\MetaTrader 5\terminal64.exe",
        ]
        for p in possiveis:
            if os.path.exists(p):
                os.startfile(p); break
        time.sleep(6)
    except Exception as e:
        print(f"Erro ao abrir MT5: {e}")

    try:
        import MetaTrader5 as mt5
        if mt5.initialize():
            tick = mt5.symbol_info_tick("XAUUSD")
            if tick:
                partes.append(f"O ouro está cotado a {tick.ask:.2f} dólares.")
    except Exception as e:
        print(f"Briefing MT5 erro: {e}")

    try:
        ticker = yf.Ticker("USDBRL=X")
        preco = ticker.fast_info['last_price']
        partes.append(f"O dólar está cotado a {preco:.2f} reais.")
    except Exception as e:
        print(f"Briefing dólar erro: {e}")

    try:
        import MetaTrader5 as mt5
        if mt5.initialize():
            ontem_inicio = agora.replace(hour=0, minute=0, second=0) - datetime.timedelta(days=1)
            ontem_fim    = ontem_inicio.replace(hour=23, minute=59, second=59)
            deals = mt5.history_deals_get(ontem_inicio, ontem_fim)
            if deals:
                lucro = round(sum(d.profit for d in deals), 2)
                if lucro >= 0:
                    partes.append(f"Ontem fechaste com lucro de {lucro:.2f} dólares.")
                else:
                    partes.append(f"Ontem fechaste com prejuízo de {abs(lucro):.2f} dólares.")
            else:
                partes.append("Não encontrei operações de ontem no MT5.")
    except Exception as e:
        print(f"Briefing trading erro: {e}")

    # Notícias — só entre 07:00 e 12:00
    if 7 <= agora.hour < 12:
        try:
            NEWS_API_KEY = "32b3a26b754f46e193ee59837b19befc"
            GROQ_API_KEY = cfg_briefing.get("Groq_Cloud")
            cliente_groq_news = Groq(api_key=GROQ_API_KEY)
            titulos_mercado = []
            titulos_brasil  = []

            # Notícias mercado financeiro — em inglês
            url_mercado = (
                f"https://newsapi.org/v2/everything?"
                f"q=gold+stock+market+forex&"
                f"language=en&sortBy=publishedAt&pageSize=5&"
                f"apiKey={NEWS_API_KEY}"
            )
            r = requests.get(url_mercado, timeout=5).json()
            for art in r.get("articles", [])[:3]:
                titulos_mercado.append(art["title"].split(" - ")[0].strip())

            # Notícias Brasil — em português
            url_brasil = (
                f"https://newsapi.org/v2/everything?"
                f"q=brasil&"
                f"language=pt&sortBy=publishedAt&pageSize=5&"
                f"apiKey={NEWS_API_KEY}"
            )
            r2 = requests.get(url_brasil, timeout=5).json()
            for art in r2.get("articles", [])[:3]:
                titulos_brasil.append(art["title"].split(" - ")[0].strip())

            # Traduz mercado via Groq
            if titulos_mercado:
                prompt = (
                    "Traduza esses títulos de notícias financeiras para português brasileiro "
                    "de forma curta e natural, um por linha, sem numeração, sem explicações:\n"
                    + "\n".join(titulos_mercado)
                )
                resp = cliente_groq_news.chat.completions.create(
                    model="llama-3.1-8b-instant",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=150
                )
                traduzidos = resp.choices[0].message.content.strip().splitlines()
                partes.append("Nos mercados mundiais:")
                for t in traduzidos:
                    if t.strip():
                        partes.append(t.strip() + ".")

            if titulos_brasil:
                partes.append("No Brasil:")
                for t in titulos_brasil:
                    partes.append(t + ".")

        except Exception as e:
            print(f"Briefing notícias erro: {e}")

    frases = [
        "Foco e disciplina fazem a diferença!",
        "O mercado recompensa quem tem paciência e estratégia!",
        "Acredita no teu método, confia no processo!",
        "Um dia de cada vez. Vamos fazer este dia valer!",
        "Consistência é a chave do sucesso no trading!",
    ]
    partes.append(random.choice(frases))

    servidor.enviar_push("🤖 VIRTUA", "Sistema inicializado")
    for parte in partes:
        falar(parte)
        time.sleep(0.3)


def modo_proativo():
    SESSOES = [
        {"nome": "Tóquio",    "hora": 0,  "minuto": 0,  "msg": lambda: f"{NOME}, a sessão de Tóquio abriu! Atenção ao mercado asiático."},
        {"nome": "Londres",   "hora": 5,  "minuto": 0,  "msg": lambda: f"{NOME}, a sessão de Londres abriu! O mercado europeu está ativo."},
        {"nome": "Nova York", "hora": 10, "minuto": 0,  "msg": lambda: f"{NOME}, a sessão de Nova York abriu! Hora de ficar de olho no ouro."},
        {"nome": "B3",        "hora": 10, "minuto": 0,  "msg": lambda: f"{NOME}, a perca de tempo da B3 abriu!"},
    ]
    sessoes_avisadas  = set()
    agenda_avisada    = set()
    ultimo_preco_ouro = None

    while True:
        try:
            agora = datetime.datetime.now()
            chave_dia = agora.strftime("%d/%m/%Y")

            for sessao in SESSOES:
                chave = f"{sessao['nome']}_{chave_dia}"
                if (agora.hour == sessao["hora"] and
                    agora.minute == sessao["minuto"] and
                    chave not in sessoes_avisadas):
                    falar(sessao["msg"]())
                    servidor.enviar_push("📊 MERCADO", f"Sessão de {sessao['nome']} abriu")
                    sessoes_avisadas.add(chave)
                    sessoes_avisadas = {k for k in sessoes_avisadas if chave_dia in k}

            try:
                agenda = carregar_agenda()
                hoje = agora.strftime("%d/%m/%Y")
                em_15min = (agora + datetime.timedelta(minutes=15)).strftime("%H:%M")
                for c in agenda:
                    chave_ag = f"{c['id']}_{chave_dia}"
                    if (c["data"] == hoje and c["hora"] == em_15min and
                        not c.get("avisado_15", False) and chave_ag not in agenda_avisada):
                        falar(f"{NOME}, atenção! Em 15 minutos tens: {c['descricao']}!")
                        servidor.enviar_push("📅 AGENDA", f"Em 15min: {c['descricao']} às {c['hora']}")
                        agenda_avisada.add(chave_ag)
            except Exception as e:
                print(f"Proativo agenda erro: {e}")

            try:
                import MetaTrader5 as mt5
                if mt5.initialize():
                    tick = mt5.symbol_info_tick("XAUUSD")
                    if tick:
                        preco_atual = tick.ask
                        if ultimo_preco_ouro is not None:
                            variacao = abs(preco_atual - ultimo_preco_ouro)
                            cfg = carregar_config()
                            limite = cfg.get('ouro_alerta', 10)
                            if variacao >= limite:
                                direcao = "subiu" if preco_atual > ultimo_preco_ouro else "caiu"
                                falar(f"{NOME}, atenção! O ouro {direcao} {variacao:.1f} dólares e está agora em {preco_atual:.2f}!")
                                servidor.enviar_push("💰 OURO", f"Ouro {direcao} {variacao:.1f} → {preco_atual:.2f}")
                                ultimo_preco_ouro = preco_atual
                        else:
                            ultimo_preco_ouro = preco_atual
            except Exception as e:
                print(f"Proativo ouro erro: {e}")

        except Exception as e:
            print(f"Erro modo proativo: {e}")

        time.sleep(60)


def loop_virtua():
    global NOME
    time.sleep(8)
    threading.Thread(target=verificar_atualizacao, daemon=True).start()
    fazer_briefing()

    while True:
        if servidor.fila_comandos:
            item = servidor.fila_comandos.pop(0)
            id_cmd = item[0]; cmd = item[1]
            resposta = processar_comando(cmd)
            servidor.respostas_virtua[id_cmd] = resposta if resposta else "Ok"
            continue

        if ui and hasattr(ui, 'mic_ativo') and not ui.mic_ativo:
            time.sleep(0.5); continue

        if ui: ui.root.after(0, ui.set_waiting)
        ativacao = ouvir_comando()

        if "virtua" in ativacao:
            cmd = ativacao.replace("virtua", "").strip()
            if cmd:
                processar_comando(cmd)
                if historico_groq:
                    while True:
                        continuacao = ouvir_comando()
                        if not continuacao:
                            break
                        if any(x in continuacao for x in ["fechar", "tchau", "até logo", "obrigado", "valeu"]):
                            break
                        if "virtua" in continuacao:
                            continuacao = continuacao.replace("virtua", "").strip()
                        processar_comando(continuacao)
            else:
                falar(f"Sim {NOME}?")
                cmd = ouvir_comando()
                if cmd:
                    processar_comando(cmd)

# ==================== PONTO DE ENTRADA ====================

if __name__ == "__main__":
    
    from pyngrok import ngrok

    config = carregar_config()
    if not config.get("nome") or not config.get("claude_key") or not config.get("Groq_Cloud") or not config.get("ngrok_token") or not config.get("cidade"):
        config = tela_setup()
        if not config:
            sys.exit()

    CLAUDE_API_KEY = config["claude_key"]
    GROQ_API_KEY   = config["Groq_Cloud"]
    NGROK_TOKEN    = config["ngrok_token"]
    NOME           = config["nome"]

    cliente_claude = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    cliente_groq   = Groq(api_key=GROQ_API_KEY)

    ui = VirtuaInterface(config)

    threading.Thread(target=loop_virtua, daemon=True).start()
    threading.Thread(target=verificar_lembretes, daemon=True).start()
    threading.Thread(target=modo_proativo, daemon=True).start()
    threading.Thread(target=lambda: servidor.start(), daemon=True).start()

    time.sleep(2)
    try:
        ngrok.set_auth_token(NGROK_TOKEN)
        url = ngrok.connect(5000)
        url_str = str(url)
        print(f"URL pública ngrok: {url_str}")
        ui.add_log(f"ngrok OK")
    except Exception as e:
        print(f"Erro ngrok: {e}")

    import phone_link
    import whatsapp
    phone_link.iniciar(falar, ouvir_comando)
    whatsapp.iniciar(falar, ouvir_comando)

    ui.run()
